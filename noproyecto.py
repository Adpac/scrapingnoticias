from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import time
import pandas as pd
from pandas import ExcelWriter
import openpyxl
from openpyxl import Workbook
import requestshtml
import requests
import lxml.html as html
from pymongo import MongoClient
client=MongoClient('localhost')
db=client['Noticias']
def set_chrome_options() -> None:
    """Sets chrome options for Selenium.
    Chrome options for headless browser is enabled.
    """
    
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_prefs = {}
    chrome_options.experimental_options["prefs"] = chrome_prefs
    chrome_prefs["profile.default_content_settings"] = {"images": 2}
    chrome_options.headless=True
    chrome_options.add_argument('--log-level=1')
    return chrome_options


def paginadinamica():
    DRIVER_PATH = 'chromedriver.exe'
    driver = webdriver.Chrome(options=set_chrome_options())
    driver.get("https://unitel.bo/tag/umsa")
    botonvermas=driver.find_element(By.XPATH,'//div[@id="vermasboton"]')
    print("procesando")
    for i in range(0,20):
        driver.execute_script("arguments[0].click();", botonvermas) #Hacemos click en el boton ver mas
        print("Generando datos: ",i)
        time.sleep(5)

    listaelementos=driver.find_elements(By.XPATH,'//h5/a')
    #print(driver.page_source)
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 42
    print("lista elementos")
    for elemento in listaelementos:
        ws.append([elemento.text, elemento.get_attribute("href")])
        print(elemento.text, elemento.get_attribute("href"))
    wb.save("mate3.xlsx")
    print("--------------------------------------")
    driver.close()
    driver.quit()




def escanearnodinamica(url="https://www.eldiario.net/portal/category/politica/page/(num)/", xpath='//div[@class="jeg_inner_content"]//article//h3/a'):
    paginicial=1
    pagfinal=50
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 42
    print("lista elementos")
    for i in range(paginicial, pagfinal):
        urlpag=url.replace("(num)",str(i))
        session = requestshtml.HTMLSession()
        print(urlpag)
        r=requests.get(urlpag)
        decode=r.content.decode('utf-8')
        parser=html.fromstring(decode)
        listaelementos=parser.xpath(xpath)
        for elemento in listaelementos:
            ws.append([elemento.text, elemento.attrib['href']])
            print(elemento.text, elemento.attrib['href'])

    wb.save("mate3.xlsx")


def consultamongo():
    colnoticias=db.noticia
    coltemas=db.temasdenoticias
    listanumtemas=list(colnoticias.find({"$text": {"$search": "buscando un titular"}}, { "score": { "$meta": "textScore" }}).sort([('score', {'$meta': 'textScore'})] ).limit(10))
    listanum=[]
    for noticia in listanumtemas:
        numtema=noticia["numerotema"]
        if not(numtema in listanum):
            listanum.append(numtema)
    print(listanum)
    print("----------listanumerotemas---------")
    listatemas=list(coltemas.find({ "numtema" :{"$in":listanum}}))
    print(listatemas)
#escanearnodinamica('https://www.eldiario.net/portal/category/economia/page/(num)/','//div[@class="jeg_block_container"]//div/h3/a')
#paginadinamica()
consultamongo()